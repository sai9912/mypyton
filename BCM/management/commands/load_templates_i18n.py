import glob
import json
import os
from collections import namedtuple

import openpyxl
from django.core.management import BaseCommand

from BCM.helpers.utils import resolve_boolean_value
from BCM.models import Language
from member_organisations.models import MemberOrganisation, ProductTemplate, ProductAttribute
from member_organisations.models import ProductPackaging
from products.models.product import Product, PackageType

IMPORT_DEBUG = False
IMG_BASE = 'products/static/products/site/wizard/proddesc'


class Command(BaseCommand):
    help = 'Command to load i18n packaging type data'
    package_type_row_fields = (
        'id', 'mo_slug', 'language', 'code', 'type', 'description'
    )

    template_row_fields = (
        'id', 'pk', 'mo_slug', 'language', 'name', 'order', 'package_level', 'attributes',
        'image_url', 'ui_label',
    )

    attribute_row_fields = (
        'id', 'pk', 'mo_slug', 'language', 'path', 'definition', 'ui_mandatory', 'ui_enabled', 'ui_read_only',
        'ui_default_callable', 'ui_field_validation_callable', 'ui_form_validation_callable',
        'ui_label', 'csv_mandatory', 'csv_default_callable', 'csv_field_validation_callable',
        'csv_form_validation_callable', 'codelist_validation',
    )
    attributes_mapping = None
    product_field_names = None

    def __init__(self, stdout=None, stderr=None, no_color=False):
        super().__init__(stdout, stderr, no_color)
        self.product_field_names = tuple([field.name for field in Product._meta.fields])

    def add_arguments(self, parser):
        parser.add_argument(
            'xlsx_path_mask',
            nargs='?',
            default='./deployment/deployment-v1-2018-03-products-dev/UI_presets_v2_templates_*',
            type=str
        )

    def get_input_file_paths(self, generic_path):
        """
        Source data is provided by path with wildcards
        """

        file_list = glob.glob(generic_path)
        file_list = [
            file_path for file_path in file_list
            if os.path.isfile(file_path)
        ]
        return file_list

    def is_mandatory_fields_set(self, source_instance, mandatory_fields):
        """
        detects if some mandatory fields are unavailable
        """

        return all([
            str(getattr(source_instance, field_name, None)).strip()
            for field_name in mandatory_fields
        ])

    @staticmethod
    def get_language_from_source(source):
        language = Language.objects.filter(slug=source.language).first()
        if language:
            return language.slug
        else:
            return None

    @staticmethod
    def get_value_or_default(value, default_value):
        return value if value else default_value

    def transform_attribute_path(self, path):
        """
        Detect if field is renamed to <field_name>_i18n and adjust path accordingly
        """

        field_name = path.split('.')[-1]
        if field_name in self.product_field_names:
            return path
        elif f'{field_name}_i18n' in self.product_field_names:
            return f'{path}_i18n'
        else:
            return None

    def save_template(self, source):
        """
        Save ProductTemplate from a source instance

        template_row_fields = (
            'id', 'pk', 'language', 'name', 'order', 'package_lavel', 'attributes', 'mo_slug',
            'image_url', 'ui_label',
        )
        """

        language = self.get_language_from_source(source)
        if not language:
            print(f'save template: Wrong language for source row: {source}')
            print('Skipping..')
            return

        member_organisation = MemberOrganisation.objects.filter(slug=source.mo_slug).first()
        attributes = [
            self.attributes_mapping.get(attribute_id)
            for attribute_id in source.attributes.split(',')
            if self.attributes_mapping.get(attribute_id)
        ]
        ui_label_i18n = json.dumps({source.language: source.ui_label})

        # toclean = ProductTemplate.objects.filter(name=source.name)
        # if toclean:
        #     toclean.delete()

        product_template = ProductTemplate(
            name=source.name,
            order=source.order,
            package_level_id=source.package_level,
            member_organisation=member_organisation,
            image_url=self.get_value_or_default(source.image_url, ''),
            ui_label_i18n=ui_label_i18n
        )
        product_template.save()
        product_template.attributes.add(*attributes)

        # if is_created:
        # print(f'ProductTemplate Instance created: {product_template}')
        self.add_en_translation(product_template)
        # else:
        #     print(f'ProductTemplate Instance updated: {product_template}')
        #     setattr(product_template, f'ui_label_{language}', source.ui_label)
        #     product_template.save()

        return product_template

    def save_attribute(self, source):
        """
        Save ProductAttribute from a source instance

        attribute_row_fields = (
            'id', 'pk', 'mo_slug', 'language', 'path', 'definition_i18n', 'ui_mandatory', 'ui_enabled', 'ui_read_only',
            'ui_default_callable', 'ui_field_validation_callable', 'ui_form_validation_callable',
            'ui_label', 'csv_mandatory', 'csv_default_callable', 'csv_field_validation_callable',
            'csv_form_validation_callable', 'codelist_validation',
        )
        """

        language = self.get_language_from_source(source)
        if not language:
            print(f'save attribute: Wrong language for source row: {source}')
            print('Skipping..')
            return

        member_organisation = MemberOrganisation.objects.filter(slug=source.mo_slug).first()
        pk = self.get_value_or_default(source.pk, None)

        ui_label = self.get_value_or_default(source.ui_label, '')
        ui_label_i18n = json.dumps({source.language: ui_label})

        definition = self.get_value_or_default(source.definition, '')
        definition_i18n = json.dumps({source.language: definition})

        field_path = self.transform_attribute_path(source.path)  # convert to i18n if required

        # toclean = ProductAttribute.objects.filter(pk=pk)
        # if toclean:
        #     toclean.delete()

        product_attribute = ProductAttribute(
            pk=pk,
            path=self.get_value_or_default(field_path, ''),
            definition_i18n=definition_i18n,
            member_organisation=member_organisation,
            ui_mandatory=resolve_boolean_value(source.ui_mandatory),
            ui_enabled=resolve_boolean_value(source.ui_enabled),
            ui_read_only=resolve_boolean_value(source.ui_read_only),
            ui_default_callable=self.get_value_or_default(source.ui_default_callable, ''),
            ui_field_validation_callable=self.get_value_or_default(source.ui_field_validation_callable, ''),
            ui_form_validation_callable=self.get_value_or_default(source.ui_form_validation_callable, ''),
            ui_label_i18n=ui_label_i18n,
            csv_mandatory=resolve_boolean_value(source.csv_mandatory),
            csv_default_callable=self.get_value_or_default(source.csv_default_callable, ''),
            csv_field_validation_callable=self.get_value_or_default(source.csv_field_validation_callable, ''),
            csv_form_validation_callable=self.get_value_or_default(source.csv_form_validation_callable, ''),
            codelist_validation=self.get_value_or_default(source.codelist_validation, '')
        )
        product_attribute.save()

        # if is_created:
        # print(f'ProductAttribute Instance created: {product_attribute}')
        self.add_en_translation(product_attribute)
        # else:
        #     print(f'ProductAttribute Instance updated: {product_attribute}')
        #     setattr(product_attribute, f'ui_label_{language}', source.ui_label)
        #     product_attribute.save()

        return product_attribute

    def save_package_type(self, source, order):
        """
        Convert named tuple to model instance
        """

        language = self.get_language_from_source(source)
        if not language:
            print(f'save_package_type; Wrong language for source row: {source}')
            print('Skipping..')
            return

        member_organisation = MemberOrganisation.objects.filter(slug=source.mo_slug).first()

        if not member_organisation:
            print(f'save_package_type: Wrong member organisation for source row: {source}')
            print('Skipping..')
            return

        if str(source.type).strip():
            type_i18n = json.dumps({
                # 'en': str(source.type).strip(),
                language: str(source.type).strip(),
            })
        else:
            print(f'save_package_type: Wrong type for source row: {source}')
            print('Skipping..')
            return

        description_i18n = json.dumps({
            language: str(source.description).strip(),
        })

        code = source.code.strip()

        pt = PackageType.objects.get(code=code)

        # toclean = ProductPackaging.objects.filter(code=code)
        # if toclean:
        #    toclean.delete()

        pp = ProductPackaging(
            code=code,
            order=order,
            member_organisation=member_organisation,
            image_url=self._get_image_url(member_organisation, code),
            ui_label_i18n=type_i18n,
            ui_description_i18n=description_i18n,
            package_type=pt
        )
        pp.save()
        self.add_en_translation(pp)

    def _get_image_url(self, mo, code):
        """
        for each packagin code lookup first
        """
        # mo_glob = "{0}/{1}/{2}.*".format(IMG_BASE, mo.slug, code) -- not used
        generic_glob = "{0}/default/{1}*.*".format(IMG_BASE, code)
        unknown_glob = "{0}/{1}*.*".format(IMG_BASE, 'unknown')
        for img_set in [glob.glob(generic_glob), glob.glob(unknown_glob)]:
            if iter:
                images = []
                for img_path in img_set:
                    images.append('/static/products' + img_path.split('products')[-1])
                images.sort()
                return ';'.join(images)
        raise Exception('no images found for pack code:' + code)

    def import_attributes(self, workbook):
        """
        Imports an xlxs file to the ProductType model
        """

        # we are able to get values by "source.mo_slug" this way
        AttributeSource = namedtuple('AttributeSource', self.attribute_row_fields)

        sheet = workbook['attribute_data']

        for index, row in enumerate(sheet.iter_rows()):
            if index == 0:
                continue

            row_values = [cell.value for cell in row]
            source = AttributeSource(*row_values)

            if self.is_mandatory_fields_set(source, ['id', 'mo_slug', 'language', 'ui_label']):
                attribute = self.save_attribute(source)
                if attribute:
                    self.attributes_mapping[str(source.id)] = attribute
            else:
                print(f'Some mandatory fields are not set for row: {source}')

    def import_templates(self, workbook):
        """
        Imports an xlxs file to the ProductType model
        """

        # we are able to get values by "source.mo_slug" this way
        TemplateSource = namedtuple('TemplateSource', self.template_row_fields)

        sheet = workbook['template_data']

        for index, row in enumerate(sheet.iter_rows()):
            if index == 0:
                continue

            row_values = [cell.value for cell in row]
            source = TemplateSource(*row_values)

            if self.is_mandatory_fields_set(source, ['id', 'ui_label']) and source.ui_label:
                self.save_template(source)
            else:
                print(f'Some mandatory fields are not set for row id: {index}, source: {source}')

    def import_package_types(self, workbook):
        """
        Imports an xlxs file to the ProductType model
        """

        # we are able to get values by "source.mo_slug" this way
        PackageTypeSource = namedtuple('PackageTypeSource', self.package_type_row_fields)

        sheet = workbook['packaging_data']

        for index, row in enumerate(sheet.iter_rows()):
            if index == 0:
                continue

            row_values = [cell.value for cell in row]
            source = PackageTypeSource(*row_values)

            if self.is_mandatory_fields_set(source, ['mo_slug', 'language', 'code', 'type']):
                self.save_package_type(source, index)
            else:
                print(f'Some mandatory fields are not set for row: {source}')

    def handle(self, *args, **options):
        file_list = self.get_input_file_paths(options['xlsx_path_mask'])

        for file_path in file_list:
            self.attributes_mapping = dict()

            print(f'\nProcessing file: {file_path.split("/")[-1]}')
            workbook = openpyxl.load_workbook(file_path)
            # 1. load attributes
            self.import_attributes(workbook)

            # 2. load template
            self.import_templates(workbook)

            # 3. load package types
            self.import_package_types(workbook)

    # functions to add a default english translation
    # english language templates need to be loaded first!

    def get_translated_field_names(self, obj):
        _translated_field_suffix = '_i18n'
        return [
            item.name for item in obj._meta.fields
            if item.name.endswith(_translated_field_suffix)
        ]

    def get_en_translation(self, a, tranlsated_attribue_name):
        if type(a) == ProductAttribute:
            for pa in type(a).objects.all().filter(path=a.path):
                if hasattr(pa, tranlsated_attribue_name):
                    return getattr(pa, tranlsated_attribue_name)
        elif type(a) == ProductTemplate:
            for ta in type(a).objects.all().filter(package_level=a.package_level):
                if hasattr(ta, tranlsated_attribue_name):
                    return getattr(ta, tranlsated_attribue_name)
        elif type(a) == ProductPackaging:
            for pa in type(a).objects.all().filter(code=a.code).filter(member_organisation__slug='gs1ie'):
                if hasattr(pa, tranlsated_attribue_name):
                    return getattr(pa, tranlsated_attribue_name)
        else:
            raise Exception('not implemented')

    def add_en_translation(self, a):
        for field in self.get_translated_field_names(a):
            # try to get en translation from some other object
            tranlsated_attribue_name = '_'.join(field.split('_')[0:-1] + ['en'])
            en_tranlsation = self.get_en_translation(a, tranlsated_attribue_name)
            if en_tranlsation:
                setattr(a, tranlsated_attribue_name, en_tranlsation)
                a.save()
                # print(getattr(a, field))
