from zipfile import BadZipFile

import openpyxl
from django.db import IntegrityError
from django.db.models import Model
from rest_framework import serializers
from rest_framework import exceptions
from rest_framework.fields import empty


class BaseUploadSerializer(serializers.Serializer):
    _upload_details = None  # upload statistics stored here
    import_file = serializers.FileField(required=True, write_only=True)

    # example what you need to specify in your subclass
    # (minimal configuration is only Meta section)
    # class Meta:
    #     model = MemberOrganisation
    #     upload_to_serializer = CompanySerializer

    def __init__(self, instance=None, data=empty, **kwargs):
        super().__init__(instance, data, **kwargs)

    def to_representation(self, instance):
        result = super().to_representation(instance)
        result['upload_details'] = self._upload_details
        return result

    def update(self, instance, validated_data):
        """
        We don't actually update member organisation instance here,
        but add company organisations, so it can be considered as "updating"
        this approach simplifies logic with rest framework
        """

        self.process_uploaded_data(validated_data['import_file'])
        return instance

    def create(self, validated_data):
        raise exceptions.MethodNotAllowed('Instances creating is not allowed here')

    def process_uploaded_data(self, import_file):
        """
        Create companies
        """

        try:
            workbook = openpyxl.load_workbook(import_file)
        except BadZipFile:
            raise exceptions.ValidationError('Can\'t open the uploaded file!')

        worksheet = workbook.worksheets[0]
        fields_map, data_start_row_id = self.get_fields_map(worksheet)

        if not fields_map:
            raise exceptions.ValidationError('Can\'t find a field names row in the uploaded file!')

        self._upload_details = {
            'success_count': 0,
            'error_count': 0,
            'errors': [],
        }

        for row in worksheet.iter_rows(row_offset=data_start_row_id):
            row_dict = {
                fields_map[col_id]: field.value
                for col_id, field in enumerate(row)
            }

            if not any(row_dict.values()):  # skip rows without any data
                continue

            try:
                result = self.create_instance(row_dict)
            except exceptions.APIException as error:
                result = {'common_error': [str(error)]}

            if isinstance(result, Model):
                self._upload_details['success_count'] += 1
            else:
                self._upload_details['error_count'] += 1
                self._upload_details['errors'].append({
                    'row_id': row[0].row,
                    'error_messages': result,
                })
        return True

    def get_fields_map(self, worksheet):
        """
        retrieve the first row as fields map
        ['uuid', 'country', 'company', ..]

        :return fields_map (list), data_start_row_id (int)
        """

        serializer = self.Meta.upload_to_serializer()
        serializer_field_names = set(serializer.fields.fields.keys())

        for row in worksheet.iter_rows():
            header_field_names = [str(cell.value).lower() for cell in row]
            fields_intersection = serializer_field_names & set(header_field_names)

            if len(fields_intersection):
                # if row contains any required fields we assume it's the true header
                data_start_row_id = row[0].row  # offset will be calculated skipping this id
                return header_field_names, data_start_row_id

        return None, None

    def get_extra_context(self, fields_data):
        """
        a hook to process "upload_to_serializer" context
        """

        return {}

    def validate_upload_fields_data(self, fields_data):
        """
        A hook to validate/process fields data,
        sometimes fields should be computed by non serializer fields
        """

        return fields_data

    def configure_required_fields(self, serializer):
        """
        Mark Meta.upload_to_required_fields as required, others as not required
        """

        required_fields = getattr(self.Meta, 'upload_to_required_fields', [])

        for field_name, field in serializer.fields.items():
            if field_name in required_fields:
                field.required = True
                field.allow_blank = False
                field.allow_null = False
            else:
                field.required = False
                field.allow_blank = True
                field.allow_null = True

        return serializer

    def create_instance(self, fields_data):
        """
        Call child serializer and create instance
        """

        fields_data = self.validate_upload_fields_data(fields_data)
        context = self.context.copy()
        context.update(self.get_extra_context(fields_data))

        serializer = self.Meta.upload_to_serializer(data=fields_data, context=context)
        serializer = self.configure_required_fields(serializer)

        if serializer.is_valid():
            try:
                result = serializer.save()
            except IntegrityError as error:
                raise exceptions.APIException('Data integrity error')
            return result
        else:
            return serializer.errors
