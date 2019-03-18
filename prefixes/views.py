import tempfile
from zipfile import ZipFile

from openpyxl import load_workbook
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect, reverse
from django.conf import settings

from core import flash, flash_get_messages, jsonify
from services import prefix_service, users_service, product_service
from .forms import PrefixActionForm, StartingNumberForm
from products.models.product import Product
from barcodes.utilities import normalize
from django.db.models import Count
from users.helpers import user_agreement_required


@user_agreement_required
def prefixes_list(request):
    current_user = request.user

    if (not current_user.profile.member_organisation or
            not current_user.profile.member_organisation.gs1_enable_advanced_dashboard):
       return redirect('profile_js')

    prefixes = prefix_service.all(user=current_user)
    suspended_prefixes = prefix_service.find_suspended(user=request.user)
    transferred_prefixes = prefix_service.find_transferred(user=request.user)

    # set products count
    result = Product.objects.values('gs1_company_prefix').annotate(count=Count('gs1_company_prefix'))
    for prefix in prefixes:
        for row in result:
            if prefix.prefix == row['gs1_company_prefix']:
                setattr(prefix, 'products', row['count'])

    '''
    result_locations = db.session.query('prefix', 'locations'). \
        from_statement(text("select locations.gs1_company_prefix as prefix, 
                                    count(*) as locations
                             from locations
                             where owner_id=%s
                             group by locations.gs1_company_prefix" % current_user.id)).all()

    # set locations count
    for prefix in prefixes:
        for row in result_locations:
            if prefix.prefix == row[0]:
                setattr(prefix, 'locations', row[1])
    '''

    if request.method == 'POST':
        form = PrefixActionForm(request.POST)
        form.fields['select_prefix'].choices = [(str(p.id), p.prefix) for p in prefixes]
        if form.is_valid():
            try:
                int_prefix_id = int(form.data['select_prefix'])
            except (ValueError, TypeError):
                flash(request, 'Your selections were not valid!', 'danger')
            else:
                prefix = prefix_service.find_item(user=current_user, id=int_prefix_id)
                if not prefix:
                    raise Http404('Prefix not found')

                if prefix.is_special == 'READ-ONLY' and form.data['prefix_action'] != 'set_this':
                    flash(request, 'Read-only prefix, please contact GS1 helpdesk.', 'danger')
                else:
                    prefix_service.make_active(user=current_user, prefix=prefix.prefix)

                    prefix_action = form.data['prefix_action']

                    # Enter a new product in selected range
                    if prefix_action == 'new_product':
                        return redirect(reverse('products:add_product') + '?prefix=' + str(prefix.prefix))

                    if prefix_action == 'new_product_js':
                        return redirect(reverse('products:add_product_js') + '?prefix=' + str(prefix.prefix))


                    # Set selected range as active and go to My Products
                    elif prefix_action == 'set_this':
                        return redirect(reverse('products:products_list'))

                    # Set starting GTIN in selected range manually
                    elif prefix_action == 'starting_gtin':
                        return redirect(reverse('prefixes:prefixes_set_starting', args=(prefix.id,)))

                    # Set starting GTIN to first available number
                    elif prefix_action == 'first_available':
                        try:
                            prefix.make_starting_from()
                        except Exception as e:
                            return render(request, 'prefixes/prefix_exhausted.html',
                                                   {'current_user': current_user, 'prefix': prefix })
                        prefix_service.save(prefix)
                        flash(request, 'Starting gtin has been set to GTIN-%s' % prefix.starting_from, 'success')
                        return redirect(reverse('prefixes:prefixes_list'))

                    # new location
                    elif prefix_action == 'new_gln':
                        return redirect(reverse('user:locations.add_location') + '?prefix=' + str(prefix.prefix))

                    elif prefix_action == 'first_available_gln':
                        pass
                        '''
                        try:
                            prefix.make_starting_from_gln()
                        except Exception,e:
                            return render_template('prefixes/prefix_exhausted.html', prefix=prefix)
                        prefix_service.save(prefix)
                        return redirect(url_for('.prefixes_list'))
                        '''

                    # Export available GTINs in this range
                    elif prefix_action == 'export_available':
                        try:
                            products = ( Product.objects.filter(owner = current_user)
                                                        .filter(gs1_company_prefix = prefix.prefix)
                                                        .order_by('gtin') )
                            prfxs = prefix.get_available_gtins(products)
                            if len(prfxs) > 0:
                                tfile = tempfile.NamedTemporaryFile(suffix='.xlsx')
                                zfile = tempfile.NamedTemporaryFile(suffix='.zip')

                                file_xlsx = load_workbook(settings.PREFIXES_EXCEL_TEMPLATE)
                                ws = file_xlsx.active
                                for index, prfx in enumerate(prfxs):
                                    _ = ws.cell(column=2, row=index + 5, value=prfx)
                                file_xlsx.save(filename=tfile.name)

                                with ZipFile(zfile.name, "w") as z:
                                    export_filename = "export_%s_available.xlsx" % (prefix.prefix,)
                                    attachment_filename = "export_%s_available.%s" % (prefix.prefix, 'zip')
                                    z.write(tfile.name, export_filename)

                                send_file = open(zfile.name, 'rb')
                                response = HttpResponse(send_file, content_type='application/zip')
                                response['Content-Disposition'] = 'attachment; filename=%s' % attachment_filename
                                return response
                            else:
                                flash(request, 'There are no available GTIN numbers for current active prefix', 'danger')
                        except Exception as e:
                            flash(request, 'Error: %s' % str(e), 'danger')
        else:
            flash(request, 'You must choose a prefix and an action!', 'danger')

    form = PrefixActionForm()
    form.fields['select_prefix'].choices = [(str(p.id), p.prefix) for p in prefixes]
    try:
        selected_prefix = int(request.POST['select_prefix'])
        prefix_service.make_active(user=request.user, prefix=selected_prefix)
    except:
        selected_prefix = prefix_service.get_active(user=request.user)

    config = {'GS1_GLN_CAPABILITY': settings.GS1_GLN_CAPABILITY}

    context = {
        'current_user': current_user,
        'config': config,
        'prefixes': prefixes,
        'suspended_prefixes': suspended_prefixes,
        'transferred_prefixes': transferred_prefixes,
        'flashed_messages': flash_get_messages(request),
        'selected_prefix': selected_prefix
    }
    return render(request, 'prefixes/prefixes_list.html', context)


@user_agreement_required
def prefixes_set_starting(request, prefix_id):
    current_user = request.user
    prefix = prefix_service.find_item(user=current_user, id=prefix_id)
    if not prefix:
        raise Http404()
    sn_length = 12 - len(prefix.prefix)
    if request.method == 'POST':
        error = 'Incorrect entry. Please enter a valid number'
        form = StartingNumberForm(request.POST)
        if form.is_valid():
            if len(form.data['starting_number']) == sn_length:
                try:
                    int(form.cleaned_data['starting_number'])
                except (ValueError, TypeError):
                    pass
                else:
                    starting_number = normalize('EAN13', prefix.prefix + form.cleaned_data['starting_number'])
                    products = product_service.filter(gtin="0" + starting_number, owner=current_user).all()
                    if len(products) == 0:
                        prefix.starting_from = starting_number
                        prefix_service.save(prefix)
                        flash(request, 'Starting gtin has been set to GTIN-%s' % prefix.starting_from, 'success')
                        return redirect(reverse('prefixes:prefixes_list'))
                    else:
                        error = 'This number is already assigned. Try another one.'
        flash(request, error, 'danger')
    form = StartingNumberForm()
    if not prefix.starting_from:
        try:
            prefix.make_starting_from()
        except:
            return render(request, 'prefixes/prefix_exhausted.html',
                                   {'current_user': current_user, 'prefix': prefix})
        prefix_service.save(prefix)
    form.data['starting_number'] = prefix.starting_from[len(prefix.prefix):12]
    context = { 'current_user': current_user,
                'prefix': prefix,
                'form': form,
                'current': prefix.starting_from,
                'sn_length': sn_length,
                'flashed_messages': flash_get_messages(request) }
    return render(request, 'prefixes/set_starting.html', context=context)


@user_agreement_required
def prefixes_ajax(request):
    if request.method != 'POST':
        raise Http404()
    prefix_id = request.POST.get('pk', None)
    if not prefix_id:
        raise Http404()
    prefix = prefix_service.find_item(user=request.user, id=prefix_id)
    if not prefix:
        raise Http404()
    prefix.description = request.POST.get('value', None)
    prefix_service.save(prefix, user=request.user)
    return jsonify(success=True)
